from __future__ import annotations

import hashlib
import json
import os
import re
import time
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from .schema import DEFAULT_SCHEMA, SurveySchema

AG_THEMES_AI = ["Accessibility", "Comfort", "Better Answers", "Speed", "Freedom", "Accuracy"]
AG_THEMES_TEACHER = [
    "Accuracy - Specificity - Delivery",
    "Experience - Understanding The Question - Clarity",
    "Verbality - Better Answers - Understandable Answers - Confidence- Human Communication- Accustomed to it - Simplicity - Relatable to Reality - Accountability - When Mistaken - Learning to Use The Tools - Encouragement",
]
AI_THEMES_NO = [
    "Cheating", "Not Achieving The Objective", "Time Must Be Spent", "Limits Creativity", "Brain",
    "Concerns", "Low Effort", "Effect on Excellence and GPA", "Effect on Skills",
    "Lowering Seriousness", "Effect on the Ability of Hard Work",
]
AI_THEMES_YES = [
    "Helpful", "Supervision", "Eases the process", "Unchangeable Reality", "Time Saver",
    "Productivity", "Minimizing Mistakes", "Undetectable", "Necessity", "Strong Assistant",
]
AI_THEMES_YES_CONDITIONAL = [
    "Necessity", "Useful Tool", "Progressivity", "Unrejectable Reality", "Undetectable", "Changing Job Market",
]

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{7,}\d)")
LONG_ID_RE = re.compile(r"\b\d{6,}\b")


def clean_text(value: object) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def redact_pii(text: str) -> str:
    text = EMAIL_RE.sub("[EMAIL]", text)
    text = PHONE_RE.sub("[PHONE]", text)
    text = LONG_ID_RE.sub("[ID]", text)
    return text


def normalize_teacher_ai_stance(value: object) -> str:
    text = clean_text(value)
    if not text:
        return "Unclear"
    if "ذكاء" in text or "الاصطناعي" in text or re.search(r"\bAI\b", text, re.I):
        return "Prefer Artificial Intelligence"
    if "المعلم" in text or "معلم" in text or "الأستاذ" in text:
        return "Prefer Teachers"
    if "كلا" in text or "حسب" in text:
        return "Both / Depends"
    return "Unclear"


def normalize_policy_stance(value: object) -> str:
    text = clean_text(value)
    if not text:
        return "Unclear"
    if any(token in text for token in ("بشرط", "شرط", "شروط", "ضوابط", "وفق")):
        return "Yes, but with Conditions"
    has_yes = "نعم" in text
    has_no = "لا" in text
    if has_yes and has_no:
        return "Yes, but with Conditions"
    if has_yes:
        return "Yes"
    if has_no:
        return "No"
    return "Unclear"


def allowed_themes(question_id: str, stance: str) -> list[str]:
    if question_id == "teacher_ai":
        if stance == "Prefer Artificial Intelligence":
            return AG_THEMES_AI.copy()
        if stance == "Prefer Teachers":
            return AG_THEMES_TEACHER.copy()
        return sorted(set(AG_THEMES_AI + AG_THEMES_TEACHER))
    if question_id == "policy":
        if stance == "No":
            return AI_THEMES_NO.copy()
        if stance == "Yes":
            return AI_THEMES_YES.copy()
        if stance == "Yes, but with Conditions":
            return AI_THEMES_YES_CONDITIONAL.copy()
        return sorted(set(AI_THEMES_NO + AI_THEMES_YES + AI_THEMES_YES_CONDITIONAL))
    raise ValueError(f"Unknown question_id: {question_id}")


def build_open_response_frame(frame: pd.DataFrame, schema: SurveySchema = DEFAULT_SCHEMA) -> pd.DataFrame:
    required = [schema.teacher_ai_preference, schema.teacher_ai_reason, schema.policy_preference, schema.policy_reason]
    missing = [column for column in required if column not in frame.columns]
    if missing:
        raise ValueError(f"Missing open-response columns: {missing}")
    rows: list[dict[str, Any]] = []
    for row_index, row in frame.iterrows():
        teacher_reason = redact_pii(clean_text(row.get(schema.teacher_ai_reason)))
        policy_reason = redact_pii(clean_text(row.get(schema.policy_reason)))
        rows.append(
            {
                "row_index": row_index,
                "question_id": "teacher_ai",
                "stance_raw": clean_text(row.get(schema.teacher_ai_preference)),
                "stance": normalize_teacher_ai_stance(row.get(schema.teacher_ai_preference)),
                "reason_text": teacher_reason,
            }
        )
        rows.append(
            {
                "row_index": row_index,
                "question_id": "policy",
                "stance_raw": clean_text(row.get(schema.policy_preference)),
                "stance": normalize_policy_stance(row.get(schema.policy_preference)),
                "reason_text": policy_reason,
            }
        )
    out = pd.DataFrame(rows)
    out["has_text"] = out["reason_text"].str.len().gt(0)
    return out


class JsonlCache:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.values: dict[str, dict[str, Any]] = {}
        if self.path.exists():
            for line in self.path.read_text(encoding="utf-8").splitlines():
                try:
                    item = json.loads(line)
                    self.values[item["key"]] = item["value"]
                except Exception:
                    continue

    def get(self, key: str) -> dict[str, Any] | None:
        return self.values.get(key)

    def put(self, key: str, value: dict[str, Any]) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.values[key] = value
        with self.path.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps({"key": key, "value": value}, ensure_ascii=False) + "\n")


def _cache_key(question_id: str, stance: str, reason_text: str) -> str:
    payload = f"{question_id}|{stance}|{reason_text}".encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def classify_response(
    client,
    *,
    model: str,
    question_id: str,
    stance: str,
    reason_text: str,
    cache: JsonlCache | None = None,
    max_retries: int = 6,
) -> dict[str, Any]:
    reason_text = redact_pii(clean_text(reason_text))
    if not reason_text:
        return {"themes": [], "confidence": 0.0}
    allowed = allowed_themes(question_id, stance)
    key = _cache_key(question_id, stance, reason_text)
    if cache and (stored := cache.get(key)) is not None:
        return stored

    schema = {
        "type": "object",
        "additionalProperties": False,
        "required": ["themes", "confidence"],
        "properties": {
            "themes": {"type": "array", "items": {"type": "string", "enum": allowed}, "uniqueItems": True},
            "confidence": {"type": "number", "minimum": 0, "maximum": 1},
        },
    }
    prompt = (
        "Label the following de-identified Arabic survey response using zero or more themes from the allowed list. "
        "Do not infer identity. Do not add themes. Code only the stated reason.\n\n"
        f"Question: {question_id}\nStance: {stance}\nAllowed themes: {allowed}\nResponse: {reason_text}"
    )
    last_error: Exception | None = None
    for attempt in range(max_retries):
        try:
            response = client.responses.create(
                model=model,
                input=[{"role": "user", "content": prompt}],
                text={"format": {"type": "json_schema", "name": "survey_theme_code", "schema": schema, "strict": True}},
            )
            value = json.loads(response.output_text)
            value["themes"] = [theme for theme in value.get("themes", []) if theme in allowed]
            value["confidence"] = float(np.clip(value.get("confidence", 0.0), 0, 1))
            if cache:
                cache.put(key, value)
            return value
        except Exception as error:  # pragma: no cover - network-dependent
            last_error = error
            time.sleep(min(8.0, 1.5 ** attempt))
    raise RuntimeError(f"OpenAI coding failed after {max_retries} attempts: {last_error}")


def code_open_responses(
    frame: pd.DataFrame,
    output_dir: str | Path,
    schema: SurveySchema = DEFAULT_SCHEMA,
    *,
    model: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    try:
        from openai import OpenAI
    except ImportError as error:  # pragma: no cover
        raise ImportError("Install the qualitative dependency: pip install -e '.[qualitative]'") from error
    if not os.getenv("OPENAI_API_KEY"):
        raise RuntimeError("OPENAI_API_KEY must be provided through the environment.")

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    model = model or os.getenv("OPENAI_MODEL", "gpt-4.1-mini")
    client = OpenAI()
    cache = JsonlCache(output_dir / "open_response_cache.jsonl")
    open_frame = build_open_response_frame(frame, schema)
    coded_rows: list[dict[str, Any]] = []
    for row in open_frame.loc[open_frame["has_text"]].to_dict(orient="records"):
        coded = classify_response(
            client,
            model=model,
            question_id=row["question_id"],
            stance=row["stance"],
            reason_text=row["reason_text"],
            cache=cache,
        )
        coded_rows.append({**row, **coded})
    coded_frame = pd.DataFrame(coded_rows)
    if coded_frame.empty:
        counts = pd.DataFrame(columns=["question_id", "stance", "theme", "count", "n_responses", "percent"])
    else:
        exploded = coded_frame.explode("themes").rename(columns={"themes": "theme"})
        exploded = exploded[exploded["theme"].notna()]
        counts = exploded.groupby(["question_id", "stance", "theme"]).size().rename("count").reset_index()
        denominators = coded_frame.groupby(["question_id", "stance"]).size().rename("n_responses").reset_index()
        counts = counts.merge(denominators, on=["question_id", "stance"], how="left")
        counts["percent"] = 100 * counts["count"] / counts["n_responses"]
    coded_frame.to_csv(output_dir / "open_responses_coded.csv", index=False, encoding="utf-8-sig")
    counts.to_csv(output_dir / "theme_counts.csv", index=False, encoding="utf-8-sig")
    return coded_frame, counts


def build_validation_workbook(coded_frame: pd.DataFrame, output_path: str | Path, n_per_question: int = 100, seed: int = 42) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    rng = np.random.default_rng(seed)
    for question_id in ("teacher_ai", "policy"):
        subset = coded_frame[coded_frame["question_id"] == question_id].copy()
        if len(subset) > n_per_question:
            subset = subset.iloc[rng.choice(len(subset), n_per_question, replace=False)]
        sheet = workbook.create_sheet(title=question_id[:31])
        themes = sorted({theme for values in subset.get("themes", []) for theme in (values if isinstance(values, list) else [])})
        headers = ["row_index", "stance", "reason_text", "model_themes"] + [f"human_{theme}" for theme in themes] + ["notes"]
        sheet.append(headers)
        for row in subset.to_dict(orient="records"):
            values = [row.get("row_index"), row.get("stance"), row.get("reason_text"), "; ".join(row.get("themes", []))]
            values += ["" for _ in themes] + [""]
            sheet.append(values)
        if themes and sheet.max_row >= 2:
            validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            sheet.add_data_validation(validation)
            start = 5
            end = 4 + len(themes)
            validation.add(f"{sheet.cell(2, start).column_letter}2:{sheet.cell(sheet.max_row, end).column_letter}{sheet.max_row}")
        sheet.freeze_panes = "A2"
    workbook.save(output_path)
